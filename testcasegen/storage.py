"""Excel persistence (Repository pattern).

One ``.xlsx`` per Jira ticket. The visible "Test Cases" sheet holds the suite;
a hidden ``_metadata`` sheet stores the requirements hash that drives the
incremental-update logic.
"""

from __future__ import annotations

import os
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .domain import TEST_CASE_FIELDS, TestCase


DISPLAY_HEADERS = [
    "Test ID", "Jira ID", "Test Title", "Description", "Steps", "Data", "Expected Result",
]
META_SHEET = "_metadata"
SHEET_NAME = "Test Cases"
COLUMN_WIDTHS = [16, 12, 32, 42, 55, 30, 52]

# Map normalized header -> canonical field, for round-trip loading.
_HEADER_TO_FIELD = {
    h.lower().replace(" ", "_"): k for h, k in zip(DISPLAY_HEADERS, TEST_CASE_FIELDS)
}
_HEADER_TO_FIELD.update({k: k for k in TEST_CASE_FIELDS})


class ExcelRepository:
    def __init__(self, output_dir: str):
        self.output_dir = output_dir
        Path(output_dir).mkdir(parents=True, exist_ok=True)

    def path_for(self, jira_id: str) -> str:
        return os.path.join(self.output_dir, f"{jira_id}.xlsx")

    def exists(self, jira_id: str) -> bool:
        return os.path.exists(self.path_for(jira_id))

    def list_jira_ids(self) -> list[str]:
        return sorted(
            p.stem for p in Path(self.output_dir).glob("*.xlsx")
            if not p.stem.startswith("_")
        )

    def load(self, jira_id: str) -> tuple[list[TestCase], str]:
        """Return (test_cases, requirements_hash). Empty if no file."""
        path = self.path_for(jira_id)
        if not os.path.exists(path):
            return [], ""

        wb = load_workbook(path, data_only=True)
        test_cases: list[TestCase] = []

        if SHEET_NAME in wb.sheetnames:
            rows = list(wb[SHEET_NAME].iter_rows(values_only=True))
            if len(rows) > 1:
                fields = [
                    _HEADER_TO_FIELD.get(
                        str(h).strip().lower().replace(" ", "_"), f"_col{i}"
                    )
                    if h is not None else f"_col{i}"
                    for i, h in enumerate(rows[0])
                ]
                for row in rows[1:]:
                    if any(c is not None for c in row):
                        raw = {
                            fields[i]: (str(v).strip() if v is not None else "")
                            for i, v in enumerate(row)
                        }
                        test_cases.append(TestCase.from_dict(raw))

        requirements_hash = ""
        if META_SHEET in wb.sheetnames:
            for row in wb[META_SHEET].iter_rows(values_only=True):
                if row and row[0] == "requirements_hash":
                    requirements_hash = str(row[1]) if row[1] else ""
                    break

        return test_cases, requirements_hash

    def save(self, jira_id: str, test_cases: list[TestCase], requirements_hash: str) -> str:
        path = self.path_for(jira_id)
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME

        hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        hdr_font = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
        alt_fill = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_i, hdr in enumerate(DISPLAY_HEADERS, 1):
            cell = ws.cell(row=1, column=col_i, value=hdr)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[1].height = 28

        for row_i, tc in enumerate(test_cases, 2):
            record = tc.to_dict()
            use_alt = row_i % 2 == 0
            for col_i, key in enumerate(TEST_CASE_FIELDS, 1):
                cell = ws.cell(row=row_i, column=col_i, value=str(record.get(key) or "").strip())
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = border
                cell.font = Font(name="Calibri", size=10)
                if use_alt:
                    cell.fill = alt_fill

        for col_i, width in enumerate(COLUMN_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(col_i)].width = width

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(TEST_CASE_FIELDS))}1"

        ws_meta = wb.create_sheet(title=META_SHEET)
        ws_meta.sheet_state = "hidden"
        ws_meta.append(["requirements_hash", requirements_hash])
        ws_meta.append(["jira_id", jira_id])
        ws_meta.append(["test_count", len(test_cases)])

        wb.save(path)
        return path
