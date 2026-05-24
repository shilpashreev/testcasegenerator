import os
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


COLUMNS = ["test_id", "jira_id", "test_title", "description", "steps", "data", "expected_result"]
DISPLAY_HEADERS = ["Test ID", "Jira ID", "Test Title", "Description", "Steps", "Data", "Expected Result"]
META_SHEET = "_metadata"

# Map display header → internal key (for round-trip load)
_HEADER_TO_KEY = {h.lower().replace(" ", "_"): k for h, k in zip(DISPLAY_HEADERS, COLUMNS)}
# Also handle exact match
_HEADER_TO_KEY.update({k: k for k in COLUMNS})


class ExcelManager:
    def __init__(self, output_dir: str):
        self.output_dir = output_dir
        Path(output_dir).mkdir(parents=True, exist_ok=True)

    def get_path(self, jira_id: str) -> str:
        return os.path.join(self.output_dir, f"{jira_id}.xlsx")

    def exists(self, jira_id: str) -> bool:
        return os.path.exists(self.get_path(jira_id))

    def list_jira_ids(self) -> list:
        return sorted(
            p.stem for p in Path(self.output_dir).glob("*.xlsx")
            if not p.stem.startswith("_")
        )

    def load_existing(self, jira_id: str) -> dict:
        path = self.get_path(jira_id)
        if not os.path.exists(path):
            return {"test_cases": [], "requirements_hash": ""}
        try:
            wb = load_workbook(path, data_only=True)
            test_cases = []

            if "Test Cases" in wb.sheetnames:
                ws = wb["Test Cases"]
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) > 1:
                    raw_headers = rows[0]
                    # Normalize: "Test ID" → "test_id", already-normalised keys pass through
                    keys = []
                    for i, h in enumerate(raw_headers):
                        if h is None:
                            keys.append(f"_col{i}")
                            continue
                        normalised = str(h).strip().lower().replace(" ", "_")
                        keys.append(_HEADER_TO_KEY.get(normalised, normalised))

                    for row in rows[1:]:
                        if any(cell is not None for cell in row):
                            tc = {keys[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)}
                            # Keep only expected columns
                            tc = {k: tc.get(k, "") for k in COLUMNS}
                            test_cases.append(tc)

            requirements_hash = ""
            if META_SHEET in wb.sheetnames:
                ws_meta = wb[META_SHEET]
                for row in ws_meta.iter_rows(values_only=True):
                    if row and row[0] == "requirements_hash":
                        requirements_hash = str(row[1]) if row[1] else ""
                        break

            return {"test_cases": test_cases, "requirements_hash": requirements_hash}
        except Exception as exc:
            return {"test_cases": [], "requirements_hash": "", "error": str(exc)}

    def save(self, jira_id: str, test_cases: list, requirements_hash: str) -> str:
        path = self.get_path(jira_id)
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"

        hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        hdr_font = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
        alt_fill = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_i, hdr in enumerate(DISPLAY_HEADERS, 1):
            cell = ws.cell(row=1, column=col_i, value=hdr)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[1].height = 28

        for row_i, tc in enumerate(test_cases, 2):
            use_alt = (row_i % 2 == 0)
            for col_i, key in enumerate(COLUMNS, 1):
                value = str(tc.get(key) or "").strip()
                cell = ws.cell(row=row_i, column=col_i, value=value)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = border
                cell.font = Font(name="Calibri", size=10)
                if use_alt:
                    cell.fill = alt_fill

        col_widths = [16, 12, 32, 42, 55, 30, 52]
        for col_i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_i)].width = width

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

        ws_meta = wb.create_sheet(title=META_SHEET)
        ws_meta.sheet_state = "hidden"
        ws_meta.append(["requirements_hash", requirements_hash])
        ws_meta.append(["jira_id", jira_id])
        ws_meta.append(["test_count", len(test_cases)])

        wb.save(path)
        return path
